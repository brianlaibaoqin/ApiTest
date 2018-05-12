import openpyxl
import Interfacetest
import time

class ExecuteTestCase:
    def execute(self,path,sheetname,reportfile):
        try:
            wb=openpyxl.load_workbook(path)#打开Excel函数的入参是文件路径
            sheet=wb.get_sheet_by_name(sheetname)
            print("打开sheet页：",sheet.title)
            #max_row=sheet.max_row
            #print(max_row)
            for i in range (2,sheet.max_row+1):
                if ("Yes"==sheet.cell(row=i,column=10).value):#获取cell单元格的值，必须加上.value
                    testcasename=sheet.cell(row=i,column=2).value.replace("\n","").replace("\r","")
                    print("执行用例：",testcasename)
                    url_ip=sheet.cell(row=i,column=3).value.replace("\n","").replace("\r","")
                    url_service=sheet.cell(row=i,column=4).value.replace("\n","").replace("\r","")
                    url=str(url_ip)+str(url_service)
                    print("请求地址：",url)
                    request_method=sheet.cell(row=i,column=5).value.replace("\n","").replace("\r","")
                    datatype=sheet.cell(row=i,column=6).value.replace("\n","").replace("\r","")
                    params=eval(sheet.cell(row=i,column=7).value.replace("\n","").replace("\r",""))
                    headers={}
                    checkpoint=sheet.cell(row=i,column=8).value.replace("\n","").replace("\r","")
                    if request_method=="GET":
                        retype="get"
                    elif request_method=="POST" and datatype=="Form":
                        retype="post_form"
                    elif request_method=="POST" and datatype=="Json":
                        retype="post_json"
                    interface=Interfacetest.InterfaceTest()#直接import模块名称的话，使用的时候必须用模块名.否则就使用from...import...
                    interface.callrequests(url,params,headers,retype,checkpoint,row=i,sheet=sheet)#方法名最好是copy
                else:
                    continue
            #reportfile=reportfile
            wb.save(reportfile)
        except Exception as e:
            print("读取用例时遇到错误：%s",str(e))

#path="C:/Users/flyyujunmh/PycharmProjects/unitest/testcase/laohuangli-testcase.xlsx"
#sheetname="TestCase"
#Exe=ExecuteTestCase()
#Exe.execute(path,sheetname)

'''
思路：
import openpyxl
dir=Excel 目录
打开Excel文件的sheet页
判断总的用例行数totalrow，和列数
for （行数）：
    for（列数）：
        sheet.cell(row=1,column=1).value.replace("\n","")
        取出每个用例的参数放入interface需要的参数列表
        调用interface

Excel复习:
import openpyxl #openpyxl只适用于后缀为xlsx的Excel文件
wb=openpyxl.load_workbook(path)     #打开excel文件
sheetlist=wb.get_sheet_names    #获取所有sheet页名称
sheet=wb.get_sheet_by_name(sheetname)   #选定指定的工作表
sheet=wb.get_active_name    #选定当前活动的工作表
sheet[A1].value #获取指定单元格的数据
sheet.cell(row=1,colunm=1).value    #获取指定单元格数据，建议的方式。如需赋值的话，直接用"="赋值
sheet.max_row   #获取当前sheet页中活动的总行数
sheet.max_column    #获取当前sheet页中活动的总列数
wb.save(path)   #保存Excel文件，取决于path是新的还是旧的
'''