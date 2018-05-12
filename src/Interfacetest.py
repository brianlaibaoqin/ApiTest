import requesttest
import json
import openpyxl
import re
class InterfaceTest:
    def callrequests(self,url,data,headers,retype,checkpoint,row,sheet):
        if retype=="get":
            params=data
            r=requesttest.requesttest()
            response=r.get(url,params,headers)
            #response_json=json.dumps(response)
            #print("类型转换",str(response))
            if(re.search(checkpoint,str(response))):
                print("get请求成功")
                try:
                    sheet.cell(row=row,column=11).value="成功"
                    sheet.cell(row=row, column=12).value = str(response)
                except Exception as e:
                    sheet.cell(row=row,column=11).value=str(e)
            else:
                print("get请求失败")
                try:
                    sheet.cell(row=row,column=11).value="失败"
                    sheet.cell(row=row, column=12).value = str(response)
                except Exception as e:
                    sheet.cell(row=row,column=11).value=str(e)
        elif retype=="post_json":
            r=requesttest.requesttest()
            response=r.post_json(url,data,headers)
            if (re.search(checkpoint, str(response))):#re.search的参数都是str类型
                print("post_json请求成功")
                try:
                    sheet.cell(row=row,column=11).value="成功"
                    sheet.cell(row=row, column=12).value = str(response)
                except Exception as e:
                    sheet.cell(row=row,column=11).value=str(e)
            else:
                print("post_json请求失败")
                try:
                    sheet.cell(row=row,column=11).value="失败"
                    sheet.cell(row=row, column=12).value = str(response)
                except Exception as e:
                    sheet.cell(row=row,column=11).value=str(e)

        elif retype=="post_form":
            r=requesttest.requesttest()
            response=r.post_form(url,data,headers)
            #response_json = json.dumps(response)
            if (re.search(checkpoint, str(response))):
                print("post_form请求成功")
                try:
                    sheet.cell(row=row,column=11).value="成功"
                    sheet.cell(row=row, column=12).value = str(response)
                except Exception as e:
                    sheet.cell(row=row,column=11).value=str(e)
            else:
                print("post_form请求失败")
                try:
                    sheet.cell(row=row,column=11).value="失败"
                    sheet.cell(row=row, column=12).value = str(response)
                except Exception as e:
                    sheet.cell(row=row,column=11).value=str(e)
        else:
            print("请求类型不存在。")

'''
#调试
url = 'http://v.juhe.cn/laohuangli/d'
params = {"key":"e711bc6362b3179f5a28de7fd3ee4ace","date":"2017-3-22"}
headers={}
retype="post_json"
checkpoint="'error_code': 0"
interface=InterfaceTest()
interface.callrequests(url,params,headers,retype,checkpoint)
'''